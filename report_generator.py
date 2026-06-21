import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime

from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

class ReportGenerator:
    def __init__(self):
        print("📊 Report Generator Inizializzato")

    # Formattazione compositore
    def _format_composer(self, comp_text):
        """
        Formatta il compositore per il report.
        """
        if not comp_text: return "Non rilevato"
        t = str(comp_text).strip().lower()
        
        invalid_terms = ["sconosciuto", "errore", "—", "-", "nessuno"]
        
        if any(x in t for x in invalid_terms):
            return "Non rilevato"
            
        return str(comp_text).strip()

    # --- 1. EXCEL (OFFICIAL - SOLO CONFERMATI E NON CANCELLATI) ---
    def generate_excel(self, playlist, metadata=None):
        wb = Workbook()
        ws = wb.active
        ws.title = "Programma Musicale"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        headers = ["N.", "TITOLO OPERA", "COMPOSITORE / AUTORE", "ARTISTA ESECUTORE"]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        row_num = 2
        index_display = 1
        
        for song in playlist:
            if song.get('is_deleted', False):
                continue
            if not song.get("confirmed", False):
                continue

            title = song.get("title", "").strip().upper()
            artist = song.get("artist", "").strip().title()
            
            raw_comp = song.get("composer", "")
            composer = self._format_composer(raw_comp).upper()

            ws.cell(row=row_num, column=1, value=index_display).alignment = center_align
            ws.cell(row=row_num, column=2, value=title).alignment = left_align
            ws.cell(row=row_num, column=3, value=composer).alignment = left_align
            ws.cell(row=row_num, column=4, value=artist).alignment = left_align

            for c in range(1, 5):
                ws.cell(row=row_num, column=c).border = thin_border

            row_num += 1
            index_display += 1

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 30

        footer_row = row_num + 2
        info_text = f"Generato automaticamente il {datetime.now().strftime('%d/%m/%Y alle %H:%M')}"
        if metadata and "artist" in metadata:
            info_text += f" - Evento: {metadata['artist']}"
        ws.cell(row=footer_row, column=2, value=info_text).font = Font(
            italic=True, size=9, color="555555"
        )

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # --- 2. PDF OFFICIAL (SOLO CONFERMATI E NON CANCELLATI) ---
    def generate_pdf_official(self, playlist, metadata=None):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        story.append(Paragraph("Programma Musicale - Borderò ufficiale", styles["Title"]))
        subtitle = "Elenco dei brani confermati."
        if metadata and "artist" in metadata:
            subtitle += f" Evento: {metadata['artist']}."
        story.append(Paragraph(subtitle, styles["Normal"]))
        story.append(Spacer(1, 12))

        header_style = ParagraphStyle(
            "HeaderStyle",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=9,
            textColor=colors.white,
            alignment=1,
        )

        data = [[
            Paragraph("N.", header_style),
            Paragraph("Titolo opera", header_style),
            Paragraph("Compositore / Autore", header_style),
            Paragraph("Artista esecutore", header_style),
        ]]

        cell_style = ParagraphStyle("CellStyle", parent=styles["Normal"], fontSize=9, leading=11)
        center_style = ParagraphStyle("CenterStyle", parent=styles["Normal"], fontSize=9, alignment=1)

        idx = 1
        for song in playlist:
            if song.get('is_deleted', False):
                continue
            if not song.get("confirmed", False):
                continue

            title = (song.get("title") or "").strip()
            artist = (song.get("artist") or "").strip()
            
            raw_comp = song.get("composer")
            composer = self._format_composer(raw_comp)

            data.append([
                Paragraph(str(idx), center_style),
                Paragraph(title, cell_style),
                Paragraph(composer, cell_style),
                Paragraph(artist, cell_style),
            ])
            idx += 1

        if len(data) == 1:
            data.append(["—", "Nessun brano confermato", "", ""])

        table = Table(data, colWidths=[30, 205, 170, 120], repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))

        story.append(table)
        story.append(Spacer(1, 12))

        info_text = f"Generato automaticamente il {datetime.now().strftime('%d/%m/%Y alle %H:%M')}"
        if metadata and "artist" in metadata:
            info_text += f" - Evento: {metadata['artist']}"
        story.append(Paragraph(info_text, styles["Italic"]))

        doc.build(story)
        buffer.seek(0)
        return buffer

    # --- 3. PDF RAW (MODIFICATO PER REQUISITI SPECIFICI) ---
    def generate_pdf_raw(self, playlist, metadata=None):
        """
        PDF 'Log Tecnico':
        - Inserimento Manuale: Riga Verde, ID="INSERIMENTO MANUALE" (No numero).
        - Rimossa: Riga Rossa, ID="{N}\nRIMOSSA". Dati Originali.
        - Modificata: Riga Nera (Dati Originali), ID="{N}\nMODIFICA AVVENUTA" (Blu).
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        story.append(Paragraph("Log di Rilevamento Automatico", styles["Title"]))
        
        subtitle_text = "Report tecnico di legittimità."
        story.append(Paragraph(subtitle_text, styles["Normal"]))
        story.append(Spacer(1, 12))

        # Stili
        h_style = ParagraphStyle("RawHeader", parent=styles["Normal"], fontName="Courier-Bold", fontSize=8, textColor=colors.white, alignment=1)
        c_style = ParagraphStyle("RawCell", parent=styles["Normal"], fontName="Helvetica", fontSize=8, leading=10, alignment=0)
        c_center = ParagraphStyle("RawCellCenter", parent=styles["Normal"], fontName="Helvetica", fontSize=8, alignment=1, leading=10)

        data = [[
            Paragraph("ID / STATO", h_style),
            Paragraph("Titolo", h_style),
            Paragraph("Compositore", h_style),
            Paragraph("Artista", h_style),
        ]]

        table_styles = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2b2b2b")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f4f4")]),
        ]

        if not playlist:
            data.append(["—", "Nessun dato", "", ""])
        else:
            for song in playlist:
                is_manual = song.get("manual", False)
                is_deleted = song.get("is_deleted", False)
                song_id_num = str(song.get("id", "?"))

                # Variabili di display e stile
                row_color = colors.black
                display_id = song_id_num
                
                # Default: mostriamo i dati ORIGINALI per dimostrare la legittimità
                # Se non c'è originale (es. manuale), usiamo il corrente
                d_title = song.get("original_title") or song.get("title", "")
                d_comp = self._format_composer(song.get("original_composer") or song.get("composer", ""))
                d_art = song.get("original_artist") or song.get("artist", "")

                # 1. CASO INSERIMENTO MANUALE (VERDE)
                if is_manual:
                    row_color = colors.HexColor("#008f00")
                    display_id = "INSERIMENTO<br/>MANUALE"
                    # Per manuale non esistono dati "originali" diversi da quelli inseriti
                    d_title = song.get("title", "")
                    d_comp = song.get("composer", "")
                    d_art = song.get("artist", "")

                # 2. CASO RIMOZIONE (ROSSO)
                elif is_deleted:
                    row_color = colors.red
                    display_id = f"{song_id_num}<br/>RIMOSSA"
                    # Qui d_title/comp sono già settati agli originali sopra

                # 3. CASO MODIFICA (ID BLU, RIGA NERA CON DATI ORIGINALI)
                else:
                    # Controllo se c'è stata modifica rispetto all'originale
                    curr_title = str(song.get("title", "")).strip().lower()
                    orig_title = str(d_title).strip().lower()
                    
                    curr_comp = str(song.get("composer", "")).strip().lower()
                    # formatto l'originale allo stesso modo per il confronto
                    orig_comp_raw = song.get("original_composer") or song.get("composer", "")
                    # nota: il confronto lo facciamo sui dati raw se possibile, o stringhe pulite
                    
                    # Se i campi correnti sono diversi dagli originali salvati
                    modified = False
                    if song.get("original_title") and song.get("original_title") != song.get("title"):
                        modified = True
                    if song.get("original_composer") and song.get("original_composer") != song.get("composer"):
                        modified = True

                    if modified:
                        # La riga resta NERA (mostra dati originali), ma l'ID ha la scritta BLU
                        display_id = f"{song_id_num}<br/><font color='blue'>MODIFICA<br/>AVVENUTA</font>"
                    
                # Creazione Paragrafi
                p_id = Paragraph(display_id, ParagraphStyle("pid", parent=c_center, textColor=row_color if not (modified and not is_deleted and not is_manual) else colors.black))
                
                # Se è modificato, l'ID label è blu (gestito nel tag font sopra), ma il testo riga deve essere nero (Originale)
                # Se è cancellato o manuale, il colore riga sovrascrive tutto.
                text_color = row_color

                p_tit = Paragraph(str(d_title), ParagraphStyle("ptit", parent=c_style, textColor=text_color))
                p_comp = Paragraph(str(d_comp), ParagraphStyle("pcomp", parent=c_style, textColor=text_color))
                p_art = Paragraph(str(d_art), ParagraphStyle("part", parent=c_style, textColor=text_color))

                data.append([p_id, p_tit, p_comp, p_art])

        col_widths = [70, 180, 160, 130]
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle(table_styles))

        story.append(table)
        story.append(Spacer(1, 12))

        info_text = f"Snapshot DB generato il {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        if metadata and "artist" in metadata:
            info_text += f" | Session: {metadata['artist']}"
        story.append(Paragraph(info_text, styles["Italic"]))

        story.append(Spacer(1, 4))
        story.append(Paragraph(
            "Questo documento certifica l'output originale del sistema.",
            styles["Italic"],
        ))

        doc.build(story)
        buffer.seek(0)
        return buffer